
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import  Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl as O
import pandas as pd

## open and load the excel file
## get the identifiers from excel file
excel_file=r"C:\Users\hp\PycharmProjects\pythonProject1\PennyDS.xlsx"
excel_sheet="Astro"
wb= O.load_workbook(excel_file)
ws=wb[excel_sheet]
identifierlist=[]
for r in range(2,11):
   Identifier= ws.cell(r,27).value
   identifierlist.append(Identifier)
print(identifierlist)

## use chromedriver which selenium use to control chrome
s= Service('C:\webdrivers\chromedriver.exe')
driver=webdriver.Chrome(service=s)
driver.implicitly_wait(15)
productset = []
techlist = []
techset = []
infolist = []
technical_list=[]


## search for each product identifier (i) and get the required info
for i in identifierlist:

        driver.get(i)
        time.sleep(10)
        page = driver.page_source
        soup = BeautifulSoup(page, 'lxml')
        ## find product's name,number,description
        try:
            name = driver.find_element(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[2]/div/h1').text
            number = driver.find_element(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[1]/div').text
            description_container = driver.find_elements(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[3]/div')
            for i in description_container:
                description = i.text
        except:
            name=''
            number=''
            description=''
        ## find all images of each product and put them in image_url
        try:
            image_url = []
            image_container = soup.findAll('div', class_='woocommerce-product-gallery__image')
            for i in image_container:
                image_ur = i.findNext()
                imgu = image_ur['href']
                image_url.append(imgu)
        except:
            image_url = []

        ## find product's downloads
        try:
            downloadlist = []

            download_item = driver.find_elements(By.CLASS_NAME, 'wcpoa_attachmentbtn')
            for download in download_item:
                download_url = download.get_attribute('href')
                downloadlist.append(download_url)
        except:
            downloadlist = []


        try:
            klist = []
            vlist = []
            specification_content = soup.findAll('div', class_="et_pb_tab clearfix")
            for m in specification_content[:1]:
                speclist = m.text.strip().split('\n')
                speclist[0]=speclist[0].lower()
                if number=="Item # 1812":
                    speclist[13] = speclist[13][:15] + ":" + speclist[13][15 + 1:]
                    speclist[14] = speclist[14].replace("-", ":", 1)
                    speclist[15] = speclist[15].replace("-", ":", 1)
                    speclist[16] = speclist[16].replace("-", ":", 1)
            if "specifications".lower() in speclist[0]:
                    del speclist[0]
            for n in speclist:
                klist.append(n.split(':')[0])
                vlist.append(n.split(':')[1])
            technicaldict = dict()
            for k in range(0, len(speclist)):
                technicaldict[klist[k]]=vlist[k]
            technical_list.append(technicaldict)


        except:

            technical_list.append({})

        print(number)
        print(technical_list)

        ## find product's info
        try:
            infoAttlist = []
            infoVallist = []
            infoValUlist = []
            info = soup.findAll('th', class_='woocommerce-product-attributes-item__label')
            for i in info:
                infoAtttext = i.text
                infoAttlist.append(infoAtttext)
            info2 = soup.findAll('td', class_='woocommerce-product-attributes-item__value')
            for n in info2:
               infoValtext = n.find('p')
               valtag = infoValtext.find('a')
               valurl = valtag['href']
               infoValUlist.append(valurl)

            infodict = dict()
            for m in range(0, len(infoValUlist)):
               infodict[infoAttlist[m]] = infoValUlist[m]
            infolist.append(infodict)
        except:
          infolist=[]

        ## create productdetails dictionary then append it to productset
        try:
            productdetails = {

                'Product Name': name,
                'Product Number': number,
                'Product Description': description,
                'Product Image': image_url,
                'Product Downloads': downloadlist


            }
            productset.append(productdetails)
        except:
            productset=[]

driver.close()

## put the product list in data frame
## save it into excel file
df1 = pd.DataFrame(productset)
df2=pd.DataFrame(infolist)
df3=pd.DataFrame(technical_list)
book=O.load_workbook(excel_file)
writer=pd.ExcelWriter("PennyDS.xlsx", engine="openpyxl")
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df1.to_excel(writer, sheet_name="Astro", startrow=0, startcol=27, index=False)
df2.to_excel(writer, sheet_name="Astro", startrow=0, startcol=32, index=False)
df3.to_excel(writer, sheet_name="Astro", startrow=0, startcol=36, index=False)
writer.save()
print('Done')
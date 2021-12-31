from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import  Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl as O
import pandas as pd

identifierlist1=[]
###read from excel file

excel_file=r"C:\Users\hp\Desktop\DataSetPenny.xlsx"
excel_sheet="Knipex"
wb= O.load_workbook(excel_file)
ws=wb[excel_sheet]
maxrow=ws.max_row

### go through the product identifier column in the excel sheet and store them in identifierlist
## for loop for all products


## for loop to test few products
for r in range(2,maxrow+1):
   Identifier= ws.cell(r,2).value
   identifierlist1.append(Identifier)
### get the main website url from excel and store it in source
Source= ws.cell(r,26).value

s= Service('C:\webdrivers\chromedriver.exe')
driver=webdriver.Chrome(service=s)
driver.implicitly_wait(15)

url = Source
driver.get(url) ## get the main url
search = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div[3]/div[2]/a[2]') ## find products button
search.send_keys(Keys.ENTER) ## press product button
## search for each identifier and get the required info
productlist = []
## list of products identifiers that are written in the dataset excel sheet with extra number which is "0"
## product 98 55 is not existed
rlist = {'11 02 0160' :'11 02 160',
'03 01 0180' : '03 01 180',
'03 02 0180' :'03 02 180',
'09 02 0240' :'09 02 240',
'03 06 0180' : '03 06 180',
'03 06 0200':'03 06 200',
'11 06 0160':'11 06 160',
  '98 55'  : ' '
         }
##check if the identifier belongs to the rlist then replace it with the correct value
def check(id):
    for n in rlist:
      if id==n:
         id= rlist[n]
    search.clear()
    search.send_keys(id)

for i in identifierlist1:
    search = driver.find_element(By.XPATH, '//*[@id="edit-search-api-fulltext--2"]') ## hit search label
    search.clear()
    search.send_keys(i)## searh for product identifier
    check(i)
    time.sleep(8)
    search.send_keys(Keys.DOWN)## choose the first option appear
    search.send_keys(Keys.ENTER)
    time.sleep(6)
    page = driver.page_source
    soup = BeautifulSoup(page, 'lxml')
    ## find all images of each product and put them in imagelist
    try:
        imagelist = []
        image_container = soup.findAll('div', {'class': "SliderProductDetailPreview slick-initialized slick-slider"})
        for i in image_container:
          for n in i.findAll('div', {'class': "slick-list draggable"}):
            img = n.findAll('div', {'class': "field__item slick-slide"})
        len_imgcontainer = len(img + image_container)

        for i in range(1, len_imgcontainer + 1):
             try:
                 xpath = f'.//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[1]/div[1]/div/div/div[{i}] / img'
                 image_element = driver.find_element(By.XPATH, xpath)
                 imageurl = image_element.get_attribute('src')
                 imagelist.append(imageurl)
             except:
                 continue
    except:
        imagelist=[]


    try:
        spareparts_container = soup.findAll('a', {'class': "Url"})
        len_sparecontaier = len(spareparts_container)

        sparelist = []
        for i in spareparts_container:
            sp_url=i['href']
            sparelist.append(sp_url)
    except:
        sparelist = []




    ## find the technical info and split them as attributes(key) in columns  and their values(value)
    try:
        tech_container = soup.find_all('div', class_='key-value-class-item')
        value = []
        key = []
        for i in tech_container:
            key.append(i.find('div', class_='key').text.replace("\n", "").strip())
        for m in tech_container:
            value.append(m.find('div', class_='value').text.replace("\n", "").strip())
        techlist = []
        techdict = dict()
        for i in range(0, len(key)):
            techdict[key[i]] = value[i]
        techlist.append(techdict)
    except:
        techlist = []

    try:
        number = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[1]/h1/div/div[1]').text
        name = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[1]/h1/div/div[2]').text
        description = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[4]').text.strip('\n').replace("\n", "-----")
        documents = driver.find_element(By.XPATH, './/*[@id="block-knipex-theme-content"]/div/article')
        documenturl = documents.find_element(By.PARTIAL_LINK_TEXT, 'Product data sheet')
        doc_url = documenturl.get_attribute('href')
    except:
        number=''
        name=''
        description=''
        doc_url=''


    print(number)
    print(sparelist)
    print(imagelist)
    try:
        product_info = {
            'Product Number': number,
            'Product Name': name,
            'Product Description': description,
            'Product Images': imagelist,
            'Product Documents': doc_url,
            'Product Spare Parts': sparelist,
            'Product Technical Details': techlist
        }
        productlist.append(product_info)
    except:
        productlist=[]
    time.sleep(15)
    driver.back()
driver.close()





identifierlist2=[]
###$$$$read from excel file$$$$

excel_file2=r"C:\Users\hp\Desktop\DataSetPenny.xlsx"
excel_sheet2="Astro"
wb2= O.load_workbook(excel_file2)
ws2=wb2[excel_sheet2]


for r in range(2,11):

   Identifier2= ws2.cell(r,27).value
   identifierlist2.append(Identifier2)
print(identifierlist2)
productset2 = []
techlist = []
techset = []
s= Service('C:\webdrivers\chromedriver.exe')
driver=webdriver.Chrome(service=s)
driver.implicitly_wait(15)

for i in identifierlist2:

        url = i
        driver.get(url)
        time.sleep(10)
        page = driver.page_source
        soup = BeautifulSoup(page, 'lxml')
        image_url = []
        try:
            name = driver.find_element(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[2]/div/h1').text
            number = driver.find_element(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[1]/div').text
            description_container = driver.find_elements(By.XPATH,'//*[@id="main-content"]/div[3]/div/div/div[2]/div[2]/div[3]/div')
            for i in description_container:
                description = i.text



            try:
                image_container = soup.findAll('div', class_='woocommerce-product-gallery__image')

                for i in image_container:
                    image_ur = i.findNext()
                    imgu = image_ur['href']
                    image_url.append(imgu)
            except:
                image_url = []
        except:
            name=''
            number=''
            description=''


        try:
            downloadlist = []

            download_item = driver.find_elements(By.CLASS_NAME, 'wcpoa_attachmentbtn')
            for download in download_item:
                download_url = download.get_attribute('href')
                downloadlist.append(download_url)
        except:
            downloadlist = []

        try:
            speclist = []


            specification_content = soup.findAll('div', class_="et_pb_tab clearfix")
            for i in specification_content:
                spec_element = i.text.strip()
                speclist.append(spec_element)
        except:
            speclist[0]=[]

        infoAttlist = []
        infoVallist = []
        infoValUlist = []
        infolist = []
        try:
            info = soup.findAll('th', class_='woocommerce-product-attributes-item__label')
            for i in info:
                infoAtttext = i.text
                infoAttlist.append(infoAtttext)
        except:
          infoAttlist = []

        try:
          info2 = soup.findAll('td', class_='woocommerce-product-attributes-item__value')
          for i in info2:
            infoValtext = i.find('p')
            valtag = infoValtext.find('a')
            valurl = valtag['href']
            infoValUlist.append(valurl)


          infodict = dict()
          for m in range(0, len(infoValUlist)):
              infodict[infoAttlist[m]] = infoValUlist[m]

          infolist.append(infodict)
        except:
          infoValUlist = []

          infolist=[]

        print(number)
        print(image_url)


        try:
            productdetails = {

                'Product Name': name,
                'Product Number': number,
                'Product Description': description,
                'Product Image': image_url,
                'Product Downloads': downloadlist,
                'Product Technical Details': speclist[0],
                'Product Info': infolist,

            }
            productset2.append(productdetails)
        except:
            productset2=[]
driver.close()








df1 = pd.DataFrame(productlist)
df2 = pd.DataFrame(productset2)
book=O.load_workbook(excel_file)
writer=pd.ExcelWriter("DataSetPenny.xlsx", engine="openpyxl")
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df1.to_excel(writer, sheet_name="Knipex", startrow=0, startcol=28, index=False)
df2.to_excel(writer, sheet_name="Astro", startrow=0, startcol=28, index=False)
writer.save()
print('Done')


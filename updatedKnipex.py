import time
import openpyxl as O
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys



#read from excel file
excel_file=r"C:\Users\hp\Desktop\PennyDS.xlsx"
excel_sheet="Knipex"
wb= O.load_workbook(excel_file)
ws=wb[excel_sheet]
maxrow=ws.max_row

#go through the product identifier column in the excel sheet and store them in identifierlist
#get the main website url from excel and store it in source
identifierlist=[]
for r in range(2,maxrow+1):
   Identifier= ws.cell(r,2).value
   identifierlist.append(Identifier)
Source= ws.cell(2,26).value

## use chromedriver which selenium use to control chrome
s= Service('C:\webdrivers\chromedriver.exe')
driver=webdriver.Chrome(service=s)
driver.implicitly_wait(15)

## get the main page url
##press products button
## go to products page
url = Source
driver.get(url)
search = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div[3]/div[2]/a[2]')
search.send_keys(Keys.ENTER)

##check if the identifier belongs to the exceptional_products keys then replace it with the correct value
def check(id):
    exceptional_products = {'11 02 0160': '11 02 160',
             '03 01 0180': '03 01 180',
             '03 02 0180': '03 02 180',
             '09 02 0240': '09 02 240',
             '03 06 0180': '03 06 180',
             '03 06 0200': '03 06 200',
             '11 06 0160': '11 06 160'}
    for n in exceptional_products.keys():
      if id==n:
         id= exceptional_products[n]
    search.clear()
    search.send_keys(id)
## create productlist for final products information
## create techlist for product technical information
productlist = []
techlist = []
## search for each product identifier (i) and get the required info
for i in identifierlist:
    ## clear the search field
    search = driver.find_element(By.XPATH, '//*[@id="edit-search-api-fulltext--2"]')
    search.clear()
    ## search for the identifiers
    if i=='98 55':
        search.send_keys(i)
        time.sleep(8)
        search.send_keys(Keys.DOWN*2)
        search.send_keys(Keys.ENTER)
    else:
        search.send_keys(i)
        check(i)
        time.sleep(8)
        search.send_keys(Keys.DOWN)
        search.send_keys(Keys.ENTER)
    ## get the page source the use lxml parser to read html page
    time.sleep(5)
    page = driver.page_source
    soup = BeautifulSoup(page, 'lxml')
    ## find all images of each product and put them in imagelist
    try:
        imagelist = []
        image_container = soup.findAll('div', {'class': "SliderProductDetailPreview slick-initialized slick-slider"})
        for i in image_container:
          for n in i.findAll('div', {'class': "slick-list draggable"}):
            img = n.findAll('div', {'class': "field__item slick-slide"})
        len_imgcontainer = len(img + image_container)
        for i in range(1, len_imgcontainer + 1):
             try:
                 xpath = f'.//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[1]/div[1]/div/div/div[{i}] / img'
                 image_element = driver.find_element(By.XPATH, xpath)
                 imageurl = image_element.get_attribute('src')
                 imagelist.append(imageurl)
             except:
                 continue
    except:
        imagelist=[]
    ## find all spareparts of each product and put them in sparelist
    try:
        spareparts_container = soup.findAll('a', {'class': "Url"})
        len_sparecontaier = len(spareparts_container)

        sparelist = []
        for i in spareparts_container:
            sp_url=i['href']
            sparelist.append(sp_url)
    except:
        sparelist = []
    ## find the technical info and split them as attributes in key list and their values in value list
    try:
        tech_container = soup.find_all('div', class_='key-value-class-item')
        value = []
        key = []
        for i in tech_container:
            key.append(i.find('div', class_='key').text.replace("\n", "").strip())
        for m in tech_container:
            value.append(m.find('div', class_='value').text.replace("\n", "").strip())
        ## put the technical keys and values in dictionary then append it to techlist
        techdict = dict()
        for i in range(0, len(key)):
            techdict[key[i]] = value[i]
        techlist.append(techdict)
    except:
        techlist = []
    ## find product's number,name.description,documents
    try:
        number = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[1]/h1/div/div[1]').text
        name = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[1]/h1/div/div[2]').text
        description = driver.find_element(By.XPATH,'//*[@id="block-knipex-theme-content"]/div/article/div/div/div[1]/div[2]/div/div[4]').text.strip('\n').replace("\n", "-----")
        documents = driver.find_element(By.XPATH, './/*[@id="block-knipex-theme-content"]/div/article')
        documenturl = documents.find_element(By.PARTIAL_LINK_TEXT, 'Product data sheet')
        doc_url = documenturl.get_attribute('href')
    except:
        number=''
        name=''
        description=''
        doc_url=''
    print(number)
    print(techlist)
    ## create products_info dictionary then append it to productlist
    try:
        product_info = {
            'Product Number': number,
            'Product Name': name,
            'Product Description': description,
            'Product Images': imagelist,
            'Product Documents': doc_url,
            'Product Spare Parts': sparelist

        }
        productlist.append(product_info)
    except:
        productlist=[]
    ## wait for 15 sec then go back to search for next identifier
    time.sleep(15)
    driver.back()
##close the chrome driever after finishing
driver.close()
## create pandas dataframe for productlist & techlist
df1 = pd.DataFrame(productlist)
df2 = pd.DataFrame(techlist)
## load penny dataset excel file & write the scraped information to Knepex sheet
book = O.load_workbook(excel_file)
writer = pd.ExcelWriter("PennyDS.xlsx", engine="openpyxl")
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df1.to_excel(writer, sheet_name="Knipex", startrow=0, startcol=27, index=False)
df2.to_excel(writer, sheet_name="Knipex", startrow=0, startcol=33, index=False)
writer.save()
print('Done')



